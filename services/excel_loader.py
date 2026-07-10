from __future__ import annotations

import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook

MAX_ANALYSIS_ROWS = 10000
PREVIEW_ROWS = 20
MAX_FORMULA_SAMPLES = 5
MAX_TABLE_PREVIEW_COLUMNS = 12


@dataclass
class ExcelSheetSummary:
    name: str
    rows: int
    cols: int
    used_range: str
    has_hidden: bool
    preview_markdown: str
    column_metadata: List[Dict[str, Any]] = field(default_factory=list)
    stats_summary: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ExcelLoadResult:
    ok: bool
    title: str
    content: str
    source_type: str
    warnings: List[str] = field(default_factory=list)
    display_text: str = ""
    metadata: Dict[str, Any] = field(default_factory=dict)

    def as_dict(self, path: Path, selected_sheet: Optional[str]) -> Dict[str, Any]:
        payload = {
            "ok": self.ok,
            "title": self.title,
            "content": self.content,
            "source_path": str(path),
            "source_type": self.source_type,
            "file_extension": path.suffix.lower(),
            "warnings": self.warnings,
            "display_text": self.display_text,
            "metadata": self.metadata,
            "selected_sheet": selected_sheet or "",
        }
        if not payload["ok"] and "error" not in payload:
            payload["error"] = "엑셀 데이터를 읽지 못했습니다."
        return payload


class ExcelLoader:
    """High-fidelity Excel/CSV analyzer for AI context generation."""

    SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}

    def load(self, file_path: Path) -> Dict[str, Any]:
        ext = file_path.suffix.lower()
        if ext not in self.SUPPORTED_EXTENSIONS:
            return {
                "ok": False,
                "error": f"지원하지 않는 엑셀 형식입니다: {ext or '(확장자 없음)'}",
                "source_path": str(file_path),
            }

        if ext == ".csv":
            result = self._load_csv(file_path)
        else:
            result = self._load_workbook(file_path)
        return result

    # ── workbook loaders ─────────────────────────────────────────────────
    def _load_workbook(self, path: Path) -> Dict[str, Any]:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            wb = load_workbook(filename=str(path), data_only=True, read_only=True)
            wb_formula = load_workbook(filename=str(path), data_only=False, read_only=True)
        sheet_names = wb.sheetnames
        if not sheet_names:
            return {
                "ok": False,
                "error": "시트가 비어 있습니다.",
                "source_path": str(path),
            }

        sheet_summaries: List[ExcelSheetSummary] = []
        global_warnings: List[str] = []
        formula_samples: List[str] = []

        for name in sheet_names:
            sheet = wb[name]
            formula_sheet = wb_formula[name]
            summary = self._summarize_sheet(path, name, sheet, formula_sheet)
            sheet_summaries.append(summary)
            formula_samples.extend(summary.metadata.get("formula_samples", []))
            global_warnings.extend(summary.warnings)

        selected = sheet_summaries[0]
        context = self._build_excel_context(
            file_name=path.name,
            sheet_summaries=sheet_summaries,
            selected_sheet=selected,
            formula_samples=formula_samples[:MAX_FORMULA_SAMPLES],
            warnings=global_warnings,
        )

        result = ExcelLoadResult(
            ok=True,
            title=path.stem,
            content=context,
            source_type="excel",
            warnings=global_warnings,
            display_text=self._build_display_text(path.name, selected),
            metadata={
                "sheet_names": [s.name for s in sheet_summaries],
                "sheet_summaries": [self._sheet_to_dict(s) for s in sheet_summaries],
            },
        )
        return result.as_dict(path, selected_sheet=selected.name)

    def _load_csv(self, path: Path) -> Dict[str, Any]:
        df = self._read_csv_with_fallback(path)
        preview = self._df_preview(df)
        column_meta = self._analyze_columns(df)
        stats_lines = self._column_stats_for_context(column_meta)
        warnings: List[str] = []
        if len(df) > MAX_ANALYSIS_ROWS:
            warnings.append(
                f"행이 {MAX_ANALYSIS_ROWS}건을 넘어 상위 데이터만 분석했습니다."
            )
        context_lines = [
            "[엑셀 파일 정보]",
            f"파일명: {path.name}",
            "파일 유형: CSV",
            f"행 수: {len(df)}",
            f"컬럼 수: {len(df.columns)}",
            "",
            "[데이터 미리보기]",
            preview or "(미리보기가 비어 있습니다)",
            "",
            "[컬럼 요약]",
        ]
        context_lines.extend(stats_lines or ["- 컬럼이 없습니다."])
        result = ExcelLoadResult(
            ok=True,
            title=path.stem,
            content="\n".join(context_lines).strip(),
            source_type="excel",
            warnings=warnings,
            display_text=f"엑셀: {path.name} / 행 {len(df)} x 열 {len(df.columns)}",
            metadata={
                "sheet_names": [path.stem],
                "sheet_summaries": [
                    {
                        "name": path.stem,
                        "rows": len(df),
                        "cols": len(df.columns),
                        "preview_markdown": preview,
                        "column_metadata": column_meta,
                    }
                ],
            },
        )
        return result.as_dict(path, selected_sheet=path.stem)

    # ── sheet analysis helpers ───────────────────────────────────────────
    def _summarize_sheet(self, path: Path, name: str, sheet, formula_sheet) -> ExcelSheetSummary:
        rows = sheet.max_row or 0
        cols = sheet.max_column or 0
        # ReadOnlyWorksheet doesn't have dimensions attribute, calculate from max_row/max_column
        if hasattr(sheet, 'dimensions') and sheet.dimensions:
            used_range = sheet.dimensions
        elif rows > 0 and cols > 0:
            from openpyxl.utils import get_column_letter
            used_range = f"A1:{get_column_letter(cols)}{rows}"
        else:
            used_range = "A1"
        has_hidden = bool(sheet.sheet_state == "hidden")

        preview_df = self._read_sheet_preview(path, name, rows)
        preview_markdown = self._df_preview(preview_df)
        column_meta = self._analyze_columns(preview_df)
        stats_summary = self._column_stats_for_context(column_meta)
        warnings: List[str] = []
        if rows > MAX_ANALYSIS_ROWS:
            warnings.append(
                f"시트 '{name}'이 {rows}행이어서 상위 {MAX_ANALYSIS_ROWS}행만 분석했습니다."
            )
        # ReadOnlyWorksheet doesn't have merged_cells attribute
        if hasattr(sheet, 'merged_cells') and sheet.merged_cells.ranges:
            warnings.append(f"시트 '{name}'에 병합된 셀이 있습니다.")
        formula_samples = self._extract_formula_samples(formula_sheet)

        return ExcelSheetSummary(
            name=name,
            rows=rows,
            cols=cols,
            used_range=used_range,
            has_hidden=has_hidden,
            preview_markdown=preview_markdown,
            column_metadata=column_meta,
            stats_summary=stats_summary,
            warnings=warnings,
            metadata={
                "formula_samples": formula_samples,
                "column_metadata": column_meta,
                "row_count": rows,
                "col_count": cols,
            },
        )

    def _read_sheet_preview(self, path: Path, sheet_name: str, total_rows: int) -> pd.DataFrame:
        rows_to_read = min(total_rows or MAX_ANALYSIS_ROWS, MAX_ANALYSIS_ROWS)
        if rows_to_read <= 0:
            return pd.DataFrame()
        return pd.read_excel(
            path,
            sheet_name=sheet_name,
            nrows=rows_to_read,
            engine="openpyxl",
        )

    def _read_csv_with_fallback(self, path: Path) -> pd.DataFrame:
        encodings = ["utf-8", "utf-8-sig", "cp949", "euc-kr", "latin-1"]
        last_error: Optional[str] = None
        for enc in encodings:
            try:
                return pd.read_csv(path, encoding=enc)
            except Exception as exc:  # noqa: BLE001
                last_error = str(exc)
                continue
        raise RuntimeError(f"CSV를 읽는 중 오류: {last_error}")

    def _analyze_columns(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        metadata: List[Dict[str, Any]] = []
        for column in df.columns:
            series = df[column]
            meta: Dict[str, Any] = {
                "name": str(column),
                "non_null": int(series.notna().sum()),
                "null_count": int(series.isna().sum()),
                "unique": int(series.nunique(dropna=True)),
                "dtype": str(series.dtype),
            }
            if pd.api.types.is_numeric_dtype(series):
                meta["type"] = "numeric"
                describe = series.describe()
                meta["stats"] = {
                    "count": float(describe.get("count", 0)),
                    "mean": float(describe.get("mean", 0)),
                    "min": float(describe.get("min", 0)),
                    "max": float(describe.get("max", 0)),
                    "std": float(describe.get("std", 0)),
                }
            elif pd.api.types.is_datetime64_any_dtype(series):
                meta["type"] = "datetime"
                if not series.dropna().empty:
                    meta["stats"] = {
                        "min": str(series.min()),
                        "max": str(series.max()),
                    }
            else:
                meta["type"] = "text"
                top = series.dropna().value_counts().head(3)
                meta["top_values"] = top.to_dict()
            metadata.append(meta)
        return metadata

    def _column_stats_for_context(self, columns: List[Dict[str, Any]]) -> List[str]:
        lines: List[str] = []
        for meta in columns:
            base = f"- {meta['name']}"
            if meta["type"] == "numeric":
                stats = meta.get("stats", {})
                line = (
                    f"{base}: 숫자형 / 평균 {stats.get('mean', 0):.2f} / 최소 {stats.get('min', 0):.2f} / 최대 {stats.get('max', 0):.2f}"
                )
            elif meta["type"] == "datetime":
                stats = meta.get("stats", {})
                line = f"{base}: 날짜형 / {stats.get('min', '')} ~ {stats.get('max', '')}"
            else:
                top_values = meta.get("top_values", {})
                top_str = ", ".join(f"{k}({v})" for k, v in top_values.items()) if top_values else "-"
                line = f"{base}: 텍스트형 / 상위값 {top_str}"
            lines.append(line)
        return lines

    def _df_preview(self, df: pd.DataFrame | None) -> str:
        if df is None or df.empty:
            return ""
        limited = df.head(PREVIEW_ROWS)
        if limited.shape[1] > MAX_TABLE_PREVIEW_COLUMNS:
            limited = limited.iloc[:, :MAX_TABLE_PREVIEW_COLUMNS]
        return limited.to_markdown(index=False)

    def _extract_formula_samples(self, sheet) -> List[str]:
        samples: List[str] = []
        if sheet is None:
            return samples
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                if cell.data_type == "f" and cell.value:
                    samples.append(f"{cell.coordinate}: {cell.value}")
                if len(samples) >= MAX_FORMULA_SAMPLES:
                    return samples
        return samples

    def _build_excel_context(
        self,
        file_name: str,
        sheet_summaries: List[ExcelSheetSummary],
        selected_sheet: ExcelSheetSummary,
        formula_samples: List[str],
        warnings: List[str],
    ) -> str:
        header = [
            "[엑셀 파일 정보]",
            f"파일명: {file_name}",
            f"시트 수: {len(sheet_summaries)}",
            "시트 목록: " + ", ".join(s.name for s in sheet_summaries),
            "",
            f"[선택 시트] {selected_sheet.name}",
            f"크기: {selected_sheet.rows}행 x {selected_sheet.cols}열",
            f"사용 범위: {selected_sheet.used_range}",
        ]
        sections = ["\n".join(header), "", "[데이터 미리보기]", selected_sheet.preview_markdown or "(미리보기 없음)"]
        sections.extend(["", "[컬럼 요약]"])
        sections.extend(selected_sheet.stats_summary or ["- 컬럼 요약 없음"])
        if formula_samples:
            sections.extend(["", "[수식 정보]", "\n".join(f"- {s}" for s in formula_samples)])
        if warnings:
            sections.extend(["", "[주의사항]", "\n".join(f"- {w}" for w in warnings)])
        return "\n".join(sections).strip()

    def _build_display_text(self, file_name: str, sheet: ExcelSheetSummary) -> str:
        return (
            f"엑셀: {file_name} / {sheet.name} / {sheet.rows}행 x {sheet.cols}열"
        )

    def _sheet_to_dict(self, summary: ExcelSheetSummary) -> Dict[str, Any]:
        return {
            "name": summary.name,
            "rows": summary.rows,
            "cols": summary.cols,
            "used_range": summary.used_range,
            "has_hidden": summary.has_hidden,
            "preview_markdown": summary.preview_markdown,
            "column_metadata": summary.column_metadata,
            "stats_summary": summary.stats_summary,
            "warnings": summary.warnings,
            "metadata": summary.metadata,
        }
